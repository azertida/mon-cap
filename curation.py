"""Couche curée de Mon cap.

Principe arrêté avec Ana :
  « Le curé l'emporte sur OSM — mais seulement là où il affirme quelque chose.
    Case vide = pas d'avis, OSM garde sa valeur. »

Le classeur (donnees-mon-cap.xlsx) est lu directement : pas d'export CSV à
refaire à chaque modification, donc pas de risque de servir des données périmées.
"""

import re
from datetime import date, datetime
from pathlib import Path

CURATED_FILE = Path(__file__).with_name("donnees-mon-cap.xlsx")
SHEET_ACTIVE = "Lieux à ajouter"
SHEET_INACTIVE = "Abris inactifs"
NOM = "Nom d'affichage"

# Catégories de la feuille → clés internes
CAT_MAP = {
    "Pharmacie": "pharmacy",
    "Station de métro": "metro",
    "Bibliothèque": "library",
    "Centre communautaire": "community_centre",
    "Lieu de culte": "place_of_worship",
    "Centre administratif": "townhall",
    "Shopping": "mall",
    "Galerie commerçante": "mall",
}
CAT_LABEL = {
    "pharmacy": "Pharmacie", "metro": "Station de métro", "library": "Bibliothèque",
    "community_centre": "Centre communautaire", "place_of_worship": "Lieu de culte",
    "townhall": "Bâtiment communal", "mall": "Galerie commerçante",
}
DYNAMIC_OFFHOURS = {"pharmacy"}          # garde tournante


def _txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _tri_state(v):
    """Oui / Limité / Non → yes / limited / no. Vide → None (pas d'avis)."""
    s = _txt(v)
    if not s:
        return None
    s = s.lower()
    if s.startswith(("oui", "yes")):
        return "yes"
    if s.startswith(("lim", "part")):
        return "limited"
    if s.startswith(("non", "no")):
        return "no"
    return None


def _bool_or_none(v):
    s = _txt(v)
    if not s:
        return None
    s = s.lower()
    if s.startswith(("oui", "yes")):
        return True
    if s.startswith(("non", "no")):
        return False
    return None


def _as_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()[:10]
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def _split_ph(s):
    """Sépare les règles jours fériés du reste de l'horaire."""
    if not s:
        return None, False
    parts = [p.strip() for p in s.split(";") if p.strip()]
    ph = any(re.match(r"^(PH|SH)\b", p) for p in parts)
    rest = "; ".join(p for p in parts if not re.match(r"^(PH|SH)\b", p))
    return (rest or None), ph


def merge_hours(habituel, ete):
    """Assemble horaire habituel + exception estivale en une seule règle OSM.

    L'ordre est critique : en syntaxe OSM, les règles écrites EN DERNIER
    écrasent les précédentes. L'exception estivale doit donc suivre l'horaire
    habituel, et « PH off » rester en toute fin — sinon un férié de juillet
    serait rouvert par la règle estivale.
    """
    h, ph1 = _split_ph(_txt(habituel))
    e, ph2 = _split_ph(_txt(ete))

    if h and e:
        # « Jul-Aug off » d'abord : sans lui, un jour non mentionné par
        # l'horaire d'été (ex. le vendredi) resterait ouvert selon l'horaire
        # habituel. On ferme la période, puis on la redéfinit entièrement.
        saison = "; ".join("Jul-Aug " + p.strip() for p in e.split(";") if p.strip())
        out = f"{h}; Jul-Aug off; {saison}"
    elif h:
        out = h
    elif e:
        # Été seul : pas de sélecteur de mois, c'est valid_until qui borne
        # la validité (sinon on annoncerait « fermé » le reste de l'année).
        out = e
    else:
        return None

    if ph1 or ph2:
        out += "; PH off"
    return out


def _read_sheet(ws):
    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        return []
    H = {c.value: i for i, c in enumerate(rows[0]) if c.value}
    out = []
    for r in rows[1:]:
        def cell(col):
            i = H.get(col)
            return r[i].value if i is not None and i < len(r) else None
        if not _txt(cell(NOM)):
            continue
        out.append({
            "osm_id": _txt(cell("osm_id")),
            "lat": cell("lat"), "lon": cell("lon"),
            "name": _txt(cell(NOM)),
            "address": _txt(cell("Adresse")),
            "category": _txt(cell("Catégorie")),
            "hours": _txt(cell("Horaires")),
            "hours_summer": _txt(cell("Horaires été")),
            "valid_until": _as_date(cell("valid_until")),
            "wheelchair": _tri_state(cell("PMR")),
            "toilets": _bool_or_none(cell("Toilettes")),
            "presence_humaine": _bool_or_none(cell("presence_humaine")),
        })
    return out


def load(today=None):
    """Renvoie (curés_par_osm_id, ajouts_purs, ids_à_écarter, avertissements)."""
    today = today or date.today()
    warns = []
    if not CURATED_FILE.exists():
        return {}, [], set(), [f"{CURATED_FILE.name} absent : récolte OSM brute."]

    from openpyxl import load_workbook
    wb = load_workbook(CURATED_FILE, data_only=True)

    ecartes = set()
    if SHEET_INACTIVE in wb.sheetnames:
        for row in _read_sheet(wb[SHEET_INACTIVE]):
            if row["osm_id"]:
                ecartes.add(row["osm_id"])

    par_id, ajouts = {}, []
    if SHEET_ACTIVE not in wb.sheetnames:
        return {}, [], ecartes, [f"Onglet « {SHEET_ACTIVE} » introuvable."]

    for row in _read_sheet(wb[SHEET_ACTIVE]):
        # horaire fusionné, puis péremption éventuelle
        row["opening_hours"] = merge_hours(row["hours"], row["hours_summer"])
        if row["valid_until"] and today > row["valid_until"]:
            if row["opening_hours"]:
                warns.append(f"{row['name']} : horaire périmé le "
                             f"{row['valid_until']} → non renseigné")
            row["opening_hours"] = None
        elif row["hours_summer"] and not row["hours"] and not row["valid_until"]:
            warns.append(f"{row['name']} : horaire d'été sans horaire habituel "
                         f"ni valid_until — il sera servi toute l'année")

        if row["osm_id"]:
            par_id[row["osm_id"]] = row
        else:
            if row["lat"] is None or row["lon"] is None:
                warns.append(f"{row['name']} : ajout sans lat/lon — ignoré")
                continue
            ajouts.append(row)

    return par_id, ajouts, ecartes, warns


def overlay(abri, cur):
    """Applique le curé sur un abri OSM : seuls les champs affirmés écrasent."""
    for src, dst in (("name", "name"), ("address", "address")):
        if cur[src]:
            abri[dst] = cur[src]
    if cur["opening_hours"]:
        abri["opening_hours"] = cur["opening_hours"]
    for f in ("wheelchair", "toilets", "presence_humaine"):
        if cur[f] is not None:
            abri[f] = cur[f]
    if cur["lat"] is not None and cur["lon"] is not None:
        abri["lat"], abri["lon"] = float(cur["lat"]), float(cur["lon"])
    if cur["category"] and cur["category"] in CAT_MAP:
        key = CAT_MAP[cur["category"]]
        abri["category"] = key
        abri["label_fr"] = CAT_LABEL.get(key, cur["category"])
        abri["dynamic_offhours"] = key in DYNAMIC_OFFHOURS
    return abri


def to_abri(cur):
    """Fabrique un abri complet à partir d'une ligne d'ajout pur."""
    key = CAT_MAP.get(cur["category"] or "", "community_centre")
    return {
        "id": "cure/" + re.sub(r"[^a-z0-9]+", "-", cur["name"].lower()).strip("-"),
        "name": cur["name"],
        "category": key,
        "label_fr": CAT_LABEL.get(key, cur["category"]),
        "lat": float(cur["lat"]), "lon": float(cur["lon"]),
        "address": cur["address"],
        "opening_hours": cur["opening_hours"],
        "wheelchair": cur["wheelchair"],
        "toilets": cur["toilets"],
        "presence_humaine": cur["presence_humaine"],
        "bench": None, "drinking_water": None, "air_conditioning": None,
        "phone": None, "website": None,
        "dynamic_offhours": key in DYNAMIC_OFFHOURS,
    }
